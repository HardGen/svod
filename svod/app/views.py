
from openpyxl import Workbook
from openpyxl.worksheet.page import PrintPageSetup
import datetime
import json
from django.contrib import messages
from django.http import Http404, HttpRequest, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from .models import *
from django.core import serializers

from django.utils import timezone

from openpyxl.styles import  Alignment,  Font, Border, Side


from datetime import date
# Create your views here.
def login(request: HttpRequest):

    if(request.COOKIES.get('otd')):
        return redirect('main')

    if (request.POST):

        if (request.POST['itotd'] == 'nmb1'):
            response = redirect('main')
            response.set_cookie('otd','nmb1', max_age=60 * 120)
            return response
            

        id = request.POST['itotd']
        otd = Otd.objects.filter(idotd = id)
        
        if(len(otd) == 0):
            messages.add_message(request, messages.INFO, 'Нет такого отделения')
            return redirect('login')
        
        response = redirect('main')
        response.set_cookie('otd',otd[0].idotd, max_age=60 * 120)
        return response
    
    return render(request, 'app/login.html')
    

def main(request: HttpRequest):
    if(request.COOKIES.get('otd') == None):
        print(request.COOKIES.get('otd'))

    otd_id = request.COOKIES.get('otd')

    if (otd_id == 'nmb1'):
        context = {
            'otd': 0
        }
        return render(request, 'app/main.html', context=context)

    try:
        otd = Otd.objects.get(idotd = otd_id)
    except Otd.DoesNotExist:
        raise Http404('нет такого отделения')

    context = {
        'otd': otd
    }
    return render(request, 'app/main.html', context=context)


def logout(request: HttpRequest):
    response = redirect('login')
    response.delete_cookie('otd')
    return response


def get_otds(request: HttpRequest):
    if(request.COOKIES.get('otd') == None):
        redirect('login')
    
    otds = Otd.objects.all()

    if (len(otds) != 0):
        context = {
            'otds': otds
        }
        
        return render(request, 'app/Otd_lists.html', context)

    return render(request, 'app/Otd_lists.html')


def otd_details(request: HttpRequest, idotd : int):
    try:
        otd = get_object_or_404(Otd, pk=idotd)
    except Otd.DoesNotExist:
        raise Http404("No MyModel matches the given query.")
    return render(request, 'app/otd_details.html', context={'otd': otd})


def food_svod(request: HttpRequest):
    if(request.COOKIES.get('otd') == None):
        return redirect('login')
    

    

    otd_id = request.COOKIES.get('otd')
    if (otd_id == 'nmb1'):
        return render(request, 'app/food_svod.html')


    otd = Otd.objects.get(idotd = otd_id)
    
    context = {
        'otd': otd
    }
    return render(request, 'app/food_svod.html', context=context)


def get_svod_by_date(request: HttpRequest, year: int, month: int, day:int):
    start_datetime = timezone.datetime(year=year, month=month, day=day)
    delta_tm = timezone.timedelta(hours=23, minutes=59, seconds=00)

    # для админа
    if (request.COOKIES.get('otd') == 'nmb1'):
        svod = Food_svod.objects.filter(
            dt_svood__range = (start_datetime, start_datetime + delta_tm)
        )
        return JsonResponse({
            'length': svod.count(),
            'svod': list(svod.values())
        })

    #для отделений
    otd_id = request.COOKIES.get('otd')
    svod = Otd.objects.get(pk=otd_id)
    result = svod.food_svod_set.filter(
        dt_svood__range = (start_datetime, start_datetime + delta_tm)
    )



    if (result.count() == 0):
        return JsonResponse({
            'length': 0
    })

    return JsonResponse({
        'length': result.count(),
        'svod': list(result.values()),
    })



def update_value(request: HttpRequest):
    column = request.POST['column']
    time = request.POST['time']
    value = request.POST['value']
    otd_id = request.COOKIES.get('otd')

    dt = datetime.datetime.strptime(time,  "%Y-%m-%d %H:%M:%S")
    if (otd_id == 'nmb1'):
        id = int(request.POST['otd_id'])
        svod = Food_svod.objects.get(idotd_id=id, dt_svood__icontains=dt)
    else:
        svod = Food_svod.objects.get(idotd_id=otd_id, dt_svood__icontains=dt)
    if column == 'vsego':
        svod.vsego = value
    if column == 'child':
        svod.child = value
    if column == 'mam':
        svod.mam = value
    if column == 'mam_nofood':
        svod.mam_nofood = value
    if column == 'wow':
        svod.wow = value
    if column == 'zond':
        svod.zond = value
    if column == 'golod':
        svod.golod = value
    if column == 'diet_01_kis':
        svod.diet_01_kis = value
    if column == 'diet_01_bul':
        svod.diet_01_bul = value
    if column == 'diet_02':
        svod.diet_02 = value
    if column == 'diet_03':
        svod.diet_03 = value
    if column == 'diet_ovd':
        svod.diet_ovd = value
    if column == 'diet_shd':
        svod.diet_shd = value
    if column == 'diet_child':
        svod.diet_child = value
    if column == 'diet_vdb':
        svod.diet_vdb = value
    if column == 'diet_nbd':
        svod.diet_nbd = value
    if column == 'diet_nkd':
        svod.diet_nkd = value
    if column == 'diet_age_1_3':
        svod.diet_age_1_3 = value
    if column == 'diet_age_3_7':
        svod.diet_age_3_7 = value
    if column == 'diet_age_7_11':
        svod.diet_age_7_11 = value
    if column == 'diet_age_11_18':
        svod.diet_age_11_18 = value
    if column == 'dop':
        svod.dop = value
    
    svod.save()


    return JsonResponse({
        'otd': serializers.serialize('json', list([svod,]))
    })
    

def set_fio(request: HttpRequest):
    if (request.POST):
        fio = request.POST['fio_ms']
        svod = Food_svod.objects.get(idotd_id=request.COOKIES.get('otd'))
        svod.fio_ms = fio
        svod.save()

        return JsonResponse({
            'fio_ms': svod.fio_ms
        })



def create_new_food_svod(request: HttpRequest):
    if(request.COOKIES.get('otd') == None):
        return redirect('login')
    otd = Otd.objects.get(pk=request.COOKIES.get('otd'))
    new_food_svod = otd.food_svod_set.create()


    return JsonResponse({
        'svod': serializers.serialize('json', [new_food_svod, ])
    })



def report(request: HttpRequest, datetm:str ):

    if(request.COOKIES.get('otd') == None):
        return redirect('login')
    start_datetime = timezone.datetime.fromisoformat(datetm)
    delta_td = timezone.timedelta(hours=23, minutes=59, seconds=00)
    end_datetime = start_datetime + delta_td


    if (request.COOKIES.get('otd') != 'nmb1'):
        otd_id = Otd.objects.get(pk=request.COOKIES.get('otd'))

        svod = Food_svod.objects.get(idotd_id=otd_id, 
        dt_svood__range = (start_datetime, end_datetime)               
        )
        create_report_seven(otd_id, svod, start_datetime.date(), "СВОДКА ПИТАНИЯ")
        return JsonResponse({
            'message': 'success'
        })
    else:
        svod_for_all_otd = Food_svod.objects.filter(
            dt_svood__range  =(start_datetime, end_datetime)
        )
        create_report_for_all_otd(svod_for_all_otd, start_datetime.date())
        return JsonResponse({
            'message': 'success'
        })



    


def create_report_for_all_otd(svod, d: date):
    right_alignment = Alignment(
        horizontal='right'
    )
    font_6 = Font(
        size= 6
    )
    font_9 = Font(
        size= 9
    )

    font_7 = Font(
        size=7,
    )

    border = Border(
        bottom=Side(border_style="thin",
                    color='FF000000')
    )

    side = Side(border_style="thin",
        color='FF000000')
    
    around_Border = Border(
        left=side,
        right=side,
        bottom=side,
        top=side
    )


    left_border = Border( left=side )
    right_border = Border(right=side)
    top_border = Border(top=side)
    bottom_border = Border(bottom=side)
   
    wb = Workbook()
    ws = wb.active

    #меняем шрифт
    for i in range(1, 15):
        for j in range(1,100):
            cell = ws.cell(row=j, column=i)
            cell.font = Font(name='Arial')

    ws['R1'] = 'Форма №22-M3'
    ws['R1'].alignment = right_alignment
    ws['R2'] = 'к Инструкции по организации лечебного питания'
    ws['R2'].alignment = right_alignment
    ws['R3'] = 'в лечебно-проф.учреждениях КГБУЗ "НМБ №1"'
    ws['R3'].alignment = right_alignment


    for k in range(0, 6):
        for i in range(1, 6):
            cell = ws.cell(row=5+k, column=i)
            cell.border = border

    ws['A4'] = 'Сводка Питания'
    
    ws['A5'] = "На 7:00"
    cell = ws['E5']
    cell.value = d.isoformat()
    cell.alignment = Alignment(horizontal='right')
    cell.number_format = 'yyyy-mm-dd'
    
    ws['A6'] = 'Состоит больных'
    ws['E9'] = f"=SUM(B15:B{14 + len(svod)})"

    ws['A7'] = 'В т.ч. детей'
    #ws['E10'] = child

    ws['A8'] = 'Всего матерей по уходу'
    #ws['E11'] = mam

    ws['A9'] = 'Из них без питания'
    #ws['E12'] = mam_nofood

    ws['A10'] = 'Ветераны УВОВ'
    #ws['E13'] = wow

    ws.merge_cells('A12:A14')
    ws['A12'].value = "ОТД"
    ws['A12'].border = around_Border
    ws['A13'].border = around_Border
    ws['B12'].border = around_Border
    ws['B13'].border = around_Border
    ws['C13'].border = around_Border
    ws['D13'].border = around_Border
    ws['E13'].border = around_Border
    ws['F13'].border = around_Border
    ws['G13'].border = around_Border
    ws['H13'].border = around_Border
    ws['I13'].border = around_Border
    ws['J13'].border = around_Border
    ws['K13'].border = around_Border
    ws['L13'].border = around_Border
    ws['M13'].border = around_Border
    ws['N13'].border = around_Border

    ws['A14'].border = around_Border
    ws['B14'].border = around_Border
    ws['B14'].border = around_Border
    ws['C14'].border = around_Border
    ws['D14'].border = around_Border
    ws['E14'].border = around_Border
    ws['F14'].border = around_Border
    ws['G14'].border = around_Border
    ws['H14'].border = around_Border
    ws['I14'].border = around_Border
    ws['J14'].border = around_Border
    ws['K14'].border = around_Border
    ws['L14'].border = around_Border
    ws['M14'].border = around_Border
    ws['N14'].border = around_Border
    ws['O14'].border = around_Border
    ws['P14'].border = around_Border
    ws['Q14'].border = around_Border
    ws['R14'].border = around_Border
    ws['C12'].border = around_Border

    ws.merge_cells('B12:B14')
    ws['B12'].value = "кол-во больных"

    ws.merge_cells('C12:R12')
    ws['C12'].value = 'СТАНДАРТНЫЕ ДИЕТЫ ( СТОЛЫ )'
    ws['C12'].font = font_7


    ws.merge_cells('C13:C14')
    ws['C13'].value = 'ЗОНД'

    ws.merge_cells('D13:D14')
    ws['D13'].value = 'ГОЛОД'

    ws.merge_cells('E13:F13')
    ws['E13'].value = '01'

    ws['E14'].value =  'Кисель'
    ws['F14'].value =  'Бульон'
    
    ws.merge_cells('G13:G14')
    ws['G13'].value = '02'

    ws.merge_cells('H13:H14')
    ws['H13'].value = '03'

    ws.merge_cells('I13:I14')
    ws['I13'].value = 'ОВД'

    ws.merge_cells('J13:J14')
    ws['J13'].value = 'ЩД'

    ws.merge_cells('K13:K14')
    ws['K13'].value = 'ВБД'

    ws.merge_cells('L13:L14')
    ws['L13'].value = 'НБД'

    ws.merge_cells('M13:M14')
    ws['M13'].value = 'НКД'

    ws.merge_cells('O13:O14')
    ws['O13'].value = '3-7 лет'

    ws.merge_cells('P13:P14')
    ws['P13'].value = '7-11 лет'

    ws.merge_cells('Q13:Q14')
    ws['Q13'].value = '11-18 лет'

    ws.merge_cells('R13:R14')
    ws['R13'].value = 'ДОП'

    

    ws.merge_cells('N13:N14')
    ws['N13'].value = '1-3 лет'

    columns = 'ABCDEFGHIJKLMNOPQR'
    for i in range(0, len(svod)):
        if (i  < len(svod)):
            ws.insert_rows(15)
        item = svod[i]
        ws['A15'] = item.idotd.short_otd
        ws['B15'] = item.vsego
        ws['C15'] = item.zond
        ws['D15'] = item.golod
        ws['E15'] = item.diet_01_kis
        ws['F15'] = item.diet_01_bul
        ws['G15'] = item.diet_02
        ws['H15'] = item.diet_03
        ws['I15'] = item.diet_ovd
        ws['J15'] = item.diet_shd
        ws['K15'] = item.diet_vdb
        ws['L15'] = item.diet_nbd
        ws['M15'] = item.diet_nkd
        ws['N15'] = item.diet_age_1_3
        ws['O15'] = item.diet_age_3_7
        ws['P15'] = item.diet_age_7_11
        ws['Q15'] = item.diet_age_11_18
        ws['R15'] = item.dop

        
    ws["B36"]=f"=SUM(B15:B{14 + len(svod)})"
    ws['E6']=f"=SUM(B15:B{14 + len(svod)})"
    ws['C36']=f"=SUM(C15:C{14 + len(svod)})"
    ws['D36']=f"=SUM(D15:D{14 + len(svod)})"
    ws['E36']=f"=SUM(E15:E{14 + len(svod)})"
    ws['F36']=f"=SUM(F15:F{14 + len(svod)})"
    ws['G36']=f"=SUM(G15:G{14 + len(svod)})"
    ws['H36']=f"=SUM(H15:H{14 + len(svod)})"
    ws['I36']=f"=SUM(I15:I{14 + len(svod)})"
    ws['J36']=f"=SUM(J15:J{14 + len(svod)})"
    ws['K36']=f"=SUM(K15:K{14 + len(svod)})"
    ws['L36']=f"=SUM(L15:L{14 + len(svod)})"
    ws['M36']=f"=SUM(M15:M{14 + len(svod)})"
    ws['N36']=f"=SUM(N15:N{14 + len(svod)})"
    ws['O36']=f"=SUM(O15:O{14 + len(svod)})"
    ws['P36']=f"=SUM(P15:P{14 + len(svod)})"
    ws['Q36']=f"=SUM(Q15:Q{14 + len(svod)})"
    ws['R36']=f"=SUM(R15:R{14 + len(svod)})"

    sum  = 0
    for item in svod:
        sum =sum + item.child
    ws['E7'] = sum

    
    sum  = 0
    for item in svod:
        sum =sum + item.mam
    ws['E8'] = sum

        
    sum  = 0
    for item in svod:
        sum =sum + item.mam_nofood
    ws['E9'] = sum

    sum  = 0
    for item in svod:
        sum =sum + item.wow
    ws['E10'] = sum



    

    # #полная граница граница сводки  [A15:N18]
    for i in range(1, 19):
        for j in range(1, len(svod) + 2):
            cell = ws.cell(row=14 + j, column=i)
            cell.border = around_Border
            cell.font = font_7
            cell.alignment = Alignment(vertical='center', horizontal='center')




    #задаем ширину столбца
    for i in range(0, 18):
        columns = 'ABCDEFGHIJKLMNOPQR'
        
        row = ws.column_dimensions[columns[i]]
        row.width = 5.43 +  0.67

    for i in range(1, 4):
        ws.row_dimensions[i].height = 9

    for i in range(6, 42):
        ws.row_dimensions[i].height = 12

            #меняем шрифт
    for i in range(1, 20):
        for j in range(1,100):
            cell = ws.cell(row=j, column=i)
            cell.font = Font(name='Arial', size=7)
            
    for i in range(5, 10):
        ws[f"A{i}"].font = Font(name='Arial')

    for i in range(12, 15):
        for j in range(1, 20):
            ws.cell(row=i, column=j).alignment =  Alignment(horizontal='center', vertical='center')

    ws['A5'].font = Font(size=9)
    ws['A6'].font = Font(size=9)
    ws['A7'].font = Font(size=9)
    ws['A8'].font = Font(size=9)
    ws['A9'].font = Font(size=9)
    ws['A10'].font = Font(size=9)
    ws['N1'].font = font_6
    ws['N2'].font = font_6
    ws['N3'].font = font_6

    for  i in range(2, 19):
        ws.cell(row=36, column=i).font = Font(name='Arial', size=7, bold=True)


    ws['A4'].font = Font(size=12, name='Arial')
    ws['B12'].alignment =  Alignment( wrap_text=True, horizontal='center', vertical='center')

    filename = timezone.datetime.now().strftime("%m-%d-%Y %H.%M.%S")

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToWidth = 1

    wb.save(f"{filename}.xlsx")







def create_report_seven(otd: Otd, svod: Food_svod, d: date, title:str):
    right_alignment = Alignment(
        horizontal='right'
    )
    font_6 = Font(
        size= 6
    )
    font_9 = Font(
        size= 9
    )

    border = Border(
        bottom=Side(border_style="thin",
                    color='FF000000')
    )

    side = Side(border_style="thin",
        color='FF000000')
    
    around_Border = Border(
        left=side,
        right=side,
        bottom=side,
        top=side
    )


    left_border = Border( left=side )
    right_border = Border(right=side)
    top_border = Border(top=side)
    bottom_border = Border(bottom=side)
   
    wb = Workbook()
    ws = wb.active

    for i in range(1, 4):
        ws.cell(row=i, column=18).alignment = Alignment(horizontal='right')
        ws.cell(row=i, column=18).font = Font(size=6, name='Arial')

    ws['R1'] = "Форма №1-84"
    ws['R2'] = "к Инструкции по организации лечебного питания"
    ws['R3'] = "в лечебно-проф.учреждениях КГБУЗ 'НМБ №1'"

    ws['A4'].font = Font(name='Arial',size=12)
    ws['A4'] = "СВОДКА ПИТАНИЯ"

    ws['R6'].font = Font(name='Arial', size=14)
    ws['R6'].alignment = Alignment(horizontal='right')
    ws['R6'] = otd.otd_name

    for i in range(1, 6):
        for j in range(1, 7):
            ws.cell(row=7 + j,  column=i).border = bottom_border
            ws.cell(row=7 + j,  column=i).font = Font(name='Arial', size=9)


    ws["A8"]  = "На 07:00"
    ws["A9"]  = "Состоит больных"
    ws["A10"]  = "В т.ч. детей"
    ws["A11"]  = "Всего матерей по уходу"
    ws["A12"]  = "Из них без питания"
    ws["A13"]  = "Ветераны УВОВ"

    
    ws['E8'].alignment = Alignment(horizontal='right')
    ws['E8'].font = Font(name='Arial', size=13, bold=True)
    ws['E8'] = d.isoformat()
    ws['E9'] = svod.vsego
    ws['E10'] = svod.child
    ws['E11'] = svod.mam
    ws['E12'] = svod.mam_nofood
    ws['E13'] = svod.wow


    

    #задаем ширину столбца
    for i in range(0, 18):
        columns = 'ABCDEFGHIJKLMNOPQR'
        
        row = ws.column_dimensions[columns[i]]
        row.width = 5.43 +  0.67

    

    for i in range(1,19):
        for j in range(15, 19):
            ws.cell(row=j, column=i).border = around_Border
            ws.cell(row=j, column=i).font = Font(name='Arial', size=7 )
            ws.cell(row=j, column=i).alignment = Alignment(horizontal='center', vertical='center',
                                                             wrap_text=True)
            
    
    ws.merge_cells("A15:A17")
    ws.merge_cells("B15:B17")
    ws.merge_cells("C16:C17")
    ws.merge_cells("D16:D17")
    ws.merge_cells("C15:R15")
    ws.merge_cells("E16:F16")
    ws.merge_cells("G16:G17")
    ws.merge_cells("H16:H17")
    ws.merge_cells("I16:I17")
    ws.merge_cells("J16:J17")
    ws.merge_cells("K16:K17")
    ws.merge_cells("L16:L17")
    ws.merge_cells("M16:M17")
    ws.merge_cells("N16:N17")
    ws.merge_cells("O16:O17")
    ws.merge_cells("P16:P17")
    ws.merge_cells("Q16:Q17")
    ws.merge_cells("R16:R17")
    
    ws['A15'] = 'ОТД'
    ws['B15'] = 'кол-во больных'
    ws['C15'] = 'СТАНДАРТНЫЕ ДИЕТЫ ( СТОЛЫ )'
    ws['C16'] = 'ЗОНД'
    ws['D16'] = 'ГОЛОД'
    ws['E16'] = '01'
    ws['E17'] = 'Кисель'
    ws['F17'] = 'Бульон'
    ws['G16'] = '02'
    ws['H16'] = '03'
    ws['I16'] = 'ОВД'
    ws['J16'] = 'ЩД'
    ws['K16'] = 'ВБД'
    ws['L16'] = 'НДБ'
    ws['M16'] = 'НКД'
    ws['N16'] = '1-3 лет'
    ws['O16'] = '3-7 лет'
    ws['P16'] = '7-11 лет'
    ws['Q16'] = '11-18 лет'
    ws['R16'] = 'ДОП'

    ws['A18'] = otd.short_otd
    ws['B18'] = svod.vsego
    ws['C18'] = svod.zond
    ws['D18'] = svod.golod
    ws['E18'] = svod.diet_01_kis
    ws['F18'] = svod.diet_01_bul
    ws['G18'] = svod.diet_02
    ws['H18'] = svod.diet_03
    ws['I18'] = svod.diet_ovd
    ws['J18'] = svod.diet_shd
    ws['K18'] = svod.diet_vdb
    ws['L18'] = svod.diet_nbd
    ws['M18'] = svod.diet_nkd
    ws['N18'] = svod.diet_age_1_3
    ws['O18'] = svod.diet_age_3_7
    ws['P18'] = svod.diet_age_7_11
    ws['Q18'] = svod.diet_age_11_18
    ws['R18'] = svod.dop

    for i in range(27, 30):
        ws.cell(row=i, column=1).font = Font(name='Arial', size=9)
        ws.cell(row=i, column=5).font = Font(name='Arial', size=9, italic=True)

    ws['A27'] = 'Зав.отделением'
    ws['A28'] = 'Ст.мед.сестра'
    ws['A29'] = 'Дежурная мед.сестра'

    ws['E27'] = otd.zav_otd
    ws['E28'] = otd.msister
    ws['E29'] = svod.fio_ms





    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToWidth = 1

    for i in range(1, 4):
        ws.row_dimensions[i].height = 9

    filename = timezone.datetime.now().strftime("%m-%d-%Y %H.%M.%S")
    wb.save(f"{filename}.xlsx")




def set_full_border(cells):
    side = Side(border_style="thin",
        color='FF000000')
    
    around_Border = Border(
        left=side,
        right=side,
        bottom=side,
        top=side
    )
    for cell in cells:
        cell.border = around_Border